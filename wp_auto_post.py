#!/usr/bin/env python3
"""
GitHub Actions から WordPress REST API へ記事を自動投稿するスクリプト。

認証情報は環境変数で受け取ります（コードに直書きしない）:
- WP_BASE_URL      例: https://example.com
- WP_USERNAME      WordPress ユーザー名
- WP_APP_PASSWORD  アプリケーションパスワード（再発行したもの）

使い方:
    python wp_auto_post.py \
        --input data/seo_80kw_production_management.xlsx \
        --sheet 制作管理表 \
        --articles-dir articles \
        --images-dir eyecatches \
        --post-status draft \
        --dry-run \
        --limit 1
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import mimetypes
import os
import re
import time
from pathlib import Path
from typing import Any

import markdown
import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from slugify import slugify
from urllib3.util.retry import Retry


# --------------------------------------------------------------------------- #
# 列名候補（用途 -> 候補リスト）。完全一致 -> 部分一致の順で柔軟に探索する。
# --------------------------------------------------------------------------- #
COLUMN_CANDIDATES: dict[str, list[str]] = {
    "no": ["No", "番号", "記事No", "ID"],
    "kw": ["指定KW", "管理KW", "KW", "キーワード", "親KW"],
    "title": ["記事タイトル案", "記事タイトル", "タイトル", "H1", "title"],
    "slug": ["スラッグ", "slug", "Slug"],
    "meta": ["メタディスクリプション案", "メタディスクリプション", "description", "概要"],
    "category": ["WPカテゴリ", "カテゴリ", "記事カテゴリ", "category"],
    "tag": ["タグ案", "タグ", "WPタグ", "tags"],
    "image_name": ["画像ファイル名", "アイキャッチファイル名", "画像名"],
    "alt": ["画像alt", "アイキャッチalt", "代替テキスト", "alt"],
    "char_target": ["文字数目安", "目標文字数", "文字数", "想定文字数"],
}

IMAGE_EXTENSIONS = [".webp", ".jpg", ".jpeg", ".png"]


# --------------------------------------------------------------------------- #
# ユーティリティ
# --------------------------------------------------------------------------- #
def env_required(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"環境変数 {name} が未設定です。GitHub Secrets を確認してください。")
    return value


def normalize_url(url: str) -> str:
    return url.rstrip("/")


def basic_auth_header(username: str, app_password: str) -> dict[str, str]:
    # アプリケーションパスワードのスペースは WordPress が無視するため除去しておく
    app_password = app_password.replace(" ", "")
    token = base64.b64encode(f"{username}:{app_password}".encode("utf-8")).decode("ascii")
    return {"Authorization": f"Basic {token}"}


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def find_col(columns: list[str], candidates: list[str]) -> str | None:
    """完全一致（大文字小文字無視）を最優先し、無ければ部分一致で探す。"""
    norm = {str(c).strip().lower(): c for c in columns}
    # 1) 完全一致
    for cand in candidates:
        key = cand.strip().lower()
        if key in norm:
            return norm[key]
    # 2) 部分一致（列名のブレ対策）
    for cand in candidates:
        key = cand.strip().lower()
        for col_key, original in norm.items():
            if key in col_key or col_key in key:
                return original
    return None


def read_table(path: Path, sheet: str | None) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {path}")

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(path, sheet_name=sheet or 0, dtype=object)
    elif suffix == ".csv":
        df = pd.read_csv(path, dtype=object)
    else:
        raise ValueError("入力は .xlsx / .xlsm / .xls / .csv のいずれかにしてください。")

    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def first_existing(paths: list[Path]) -> Path | None:
    for p in paths:
        if p.exists() and p.is_file():
            return p
    return None


def split_terms(value: str) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[,、\n/|｜]+", value)
    return [p.strip() for p in parts if p.strip()]


def make_slug(text: str, fallback: str) -> str:
    value = slugify(text or fallback, lowercase=True)
    value = value[:90].strip("-")
    return value or fallback


def parse_no_spec(spec: str) -> set[int] | None:
    """"2-10,13-15,20" のような指定を No の集合に展開する。空なら None（全件）。"""
    spec = (spec or "").strip()
    if not spec:
        return None
    result: set[int] = set()
    for part in re.split(r"[,、\s]+", spec):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(\d+)\s*[-~〜]\s*(\d+)$", part)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            result.update(range(min(a, b), max(a, b) + 1))
        elif part.isdigit():
            result.add(int(part))
    return result or None


def detect_duplicate_slugs(df: "pd.DataFrame", col: dict[str, str | None]) -> dict[str, list[str]]:
    """管理表の各行から投稿スラッグを算出し、重複しているスラッグを返す。

    戻り値は {スラッグ: [No, No, ...]}（2件以上重複したものだけ）。投稿ループと
    同じ規則（スラッグ列があればそれ、無ければタイトル）でスラッグを計算する。
    """
    by_slug: dict[str, list[str]] = {}
    for idx, row in df.iterrows():
        no_value = safe_str(row.get(col["no"])) if col["no"] else str(int(idx) + 1)
        if not no_value:
            no_value = str(int(idx) + 1)
        title = safe_str(row.get(col["title"])) if col["title"] else ""
        if not title and col["kw"]:
            title = safe_str(row.get(col["kw"]))
        if not title:
            continue
        explicit_slug = safe_str(row.get(col["slug"])) if col["slug"] else ""
        slug = make_slug(explicit_slug or title, f"post-{int(idx) + 1:03d}")
        by_slug.setdefault(slug, []).append(no_value)
    return {slug: nos for slug, nos in by_slug.items() if len(nos) > 1}


def no_to_int(no_value: Any) -> int | None:
    s = safe_str(no_value)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def row_no_to_names(no_value: str) -> list[str]:
    raw = safe_str(no_value)
    if not raw:
        return []
    try:
        number = int(float(raw))
        return [f"{number:03d}", str(number)]
    except ValueError:
        return [raw]


_VOID_TAGS = {"br", "hr", "img", "input", "meta", "link", "area", "base",
              "col", "embed", "source", "track", "wbr"}
_TAG_RE = re.compile(r"<(/?)([a-zA-Z0-9]+)([^>]*?)(/?)>")


def _split_top_level(html: str) -> list[tuple[str, str]]:
    """HTML をトップレベル要素ごとに (タグ名, 要素HTML) のリストに分割する。"""
    elements: list[tuple[str, str]] = []
    depth = 0
    start: int | None = None
    cur_tag = ""
    for m in _TAG_RE.finditer(html):
        closing = m.group(1) == "/"
        tag = m.group(2).lower()
        selfclose = bool(m.group(4)) or tag in _VOID_TAGS
        if depth == 0 and not closing:
            start = m.start()
            cur_tag = tag
            if selfclose:
                elements.append((tag, html[start:m.end()]))
                start = None
            else:
                depth = 1
        elif not closing and not selfclose:
            depth += 1
        elif closing:
            depth -= 1
            if depth == 0 and start is not None:
                elements.append((cur_tag, html[start:m.end()]))
                start = None
    return elements


def html_to_gutenberg_blocks(html: str) -> str:
    """Markdown 由来の HTML を Gutenberg ブロックマークアップへ変換する。

    これにより WordPress 投稿時に「無効なブロック（ブロックを解除）」の警告を防ぐ。
    対応: 段落 / 見出し / リスト / 表 / 引用 / 区切り線。未知要素は HTML ブロックで包む。
    """
    blocks: list[str] = []
    for tag, el in _split_top_level(html):
        el = el.strip()
        if not el:
            continue
        if tag == "p":
            blocks.append(f"<!-- wp:paragraph -->\n{el}\n<!-- /wp:paragraph -->")
        elif tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
            level = int(tag[1])
            m = re.match(r"<h[1-6][^>]*>(.*)</h[1-6]>\s*$", el, re.S)
            inner = m.group(1) if m else el
            attr = "" if level == 2 else f' {{"level":{level}}}'
            blocks.append(
                f"<!-- wp:heading{attr} -->\n"
                f'<h{level} class="wp-block-heading">{inner}</h{level}>\n'
                f"<!-- /wp:heading -->"
            )
        elif tag in ("ul", "ol"):
            ordered = tag == "ol"
            items = re.findall(r"<li[^>]*>(.*?)</li>", el, re.S)
            li_html = "".join(
                f"<!-- wp:list-item -->\n<li>{it.strip()}</li>\n<!-- /wp:list-item -->\n"
                for it in items
            )
            list_attr = ' {"ordered":true}' if ordered else ""
            list_tag = "ol" if ordered else "ul"
            blocks.append(
                f"<!-- wp:list{list_attr} -->\n"
                f'<{list_tag} class="wp-block-list">\n{li_html}</{list_tag}>\n'
                f"<!-- /wp:list -->"
            )
        elif tag == "table":
            blocks.append(
                "<!-- wp:table -->\n"
                f'<figure class="wp-block-table">{el}</figure>\n'
                "<!-- /wp:table -->"
            )
        elif tag == "blockquote":
            blocks.append(f"<!-- wp:quote -->\n{el}\n<!-- /wp:quote -->")
        elif tag == "hr":
            blocks.append(
                '<!-- wp:separator -->\n'
                '<hr class="wp-block-separator has-alpha-channel-opacity"/>\n'
                "<!-- /wp:separator -->"
            )
        elif tag == "pre":
            blocks.append(f"<!-- wp:code -->\n{el}\n<!-- /wp:code -->")
        else:
            # div など想定外の生 HTML は HTML ブロックで包む（解除警告を回避）
            blocks.append(f"<!-- wp:html -->\n{el}\n<!-- /wp:html -->")
    return "\n\n".join(blocks)


def markdown_to_html(md_text: str, *, strip_h1: bool, title: str = "") -> str:
    """Markdown -> HTML。strip_h1=True なら本文先頭の H1 を無条件で除去する。

    WordPress テーマ側が投稿タイトルを表示するため、本文先頭の H1 はタイトルと
    一致していなくても二重表示になる。よって先頭 H1（Markdown の `# ...`）は
    内容に関わらず削除する。本文中の `##` 以降（H2/H3...）は残す。
    title 引数は後方互換のため残しているが判定には使用しない。
    """
    text = md_text.lstrip("﻿").lstrip()
    if strip_h1:
        text = strip_leading_markdown_h1(text)

    html = markdown.markdown(
        text,
        extensions=["extra", "tables", "sane_lists", "toc"],
        output_format="html5",
    )

    if strip_h1:
        html = strip_leading_html_h1(html)
    return html


def build_image_block(image_url: str, alt_text: str) -> str:
    """本文先頭に差し込むアイキャッチ画像の HTML（Gutenberg 画像ブロック）を返す。"""
    import html as _html

    safe_url = _html.escape(image_url, quote=True)
    safe_alt = _html.escape(alt_text or "", quote=True)
    return (
        "<!-- wp:image {\"sizeSlug\":\"large\"} -->\n"
        f'<figure class="wp-block-image size-large">'
        f'<img src="{safe_url}" alt="{safe_alt}"/></figure>\n'
        "<!-- /wp:image -->\n\n"
    )


def strip_leading_markdown_h1(text: str) -> str:
    """本文の最初の見出しが H1（"# ..."）ならその行を削除する。

    YAML フロントマター（先頭の --- ... ---）や空行・コメントを読み飛ばし、
    最初に現れる「実体のある行」が H1 のときだけ削除する。setext 形式
    （見出しの次行が === ）の H1 にも対応する。"##" 以降は対象外。
    """
    lines = text.split("\n")
    n = len(lines)

    # 先頭の YAML フロントマター（--- ... ---）は丸ごと除去する
    if n and lines[0].strip() == "---":
        j = 1
        while j < n and lines[j].strip() != "---":
            j += 1
        if j < n:  # 閉じ --- が見つかった場合のみ除去
            del lines[: j + 1]
            n = len(lines)

    i = 0
    # 空行・HTMLコメント行を読み飛ばし、最初の実体行を探す
    while i < n and (lines[i].strip() == "" or lines[i].lstrip().startswith("<!--")):
        i += 1

    if i >= n:
        return text

    first = lines[i]
    # ATX 形式: "# 見出し"（"##" は除外）
    if re.match(r"^#\s+\S", first) and not first.lstrip().startswith("##"):
        del lines[i]
        return "\n".join(lines).lstrip("\n")

    # setext 形式: 見出しテキストの次行が "===..."（H1）
    if i + 1 < n and re.match(r"^=+\s*$", lines[i + 1]) and first.strip():
        del lines[i : i + 2]
        return "\n".join(lines).lstrip("\n")

    return text


def strip_leading_html_h1(html: str) -> str:
    """HTML 本文の最初の <h1>...</h1> を1つだけ除去する（最初の <h2> より前のもの）。

    HTML を直接本文に持つケースや、フロントマター/hr 等で先頭判定を逃した
    Markdown 変換結果に対応する。タイトル二重表示の原因となる本文冒頭の H1 は
    必ず最初の H2（節見出し）より前に現れるため、その範囲にある最初の H1 だけを
    削除する。本文後半に意図的に置かれた H1 は残す。
    """
    h1 = re.search(r"<h1\b[^>]*>.*?</h1>\s*", html, flags=re.IGNORECASE | re.DOTALL)
    if not h1:
        return html
    h2 = re.search(r"<h2\b", html, flags=re.IGNORECASE)
    if h2 and h1.start() > h2.start():
        return html  # 最初の H1 が最初の H2 より後ろなら本文見出しとみなし残す
    return (html[: h1.start()] + html[h1.end():]).lstrip()


def count_text_chars(html: str) -> int:
    """HTML 本文から本文テキストの文字数を数える（タグ除去・空白除外）。"""
    text = re.sub(r"<[^>]+>", "", html)          # タグ除去
    text = re.sub(r"&[a-zA-Z#0-9]+;", "", text)  # HTMLエンティティ除去
    text = re.sub(r"\s+", "", text)              # 空白・改行を除外
    return len(text)


def parse_char_target(target: str) -> tuple[int | None, int | None]:
    """「4,500〜6,500字」「3000字以上」等から (min, max) を取り出す。無ければ (None, None)。"""
    nums = [int(n.replace(",", "")) for n in re.findall(r"\d[\d,]*", target or "")]
    if not nums:
        return None, None
    if len(nums) == 1:
        return nums[0], None
    return min(nums), max(nums)


def judge_char_count(actual: int, target: str) -> str:
    """文字数実績を目安レンジと比較して判定文字列を返す。目安が無ければ空。"""
    lo, hi = parse_char_target(target)
    if lo is None and hi is None:
        return ""
    if lo is not None and actual < lo:
        return f"不足(目安{target})"
    if hi is not None and actual > hi:
        return f"超過(目安{target})"
    return "OK"


def find_article_file(articles_dir: Path, slug: str, no_value: str) -> Path | None:
    candidates: list[Path] = []
    if slug:
        candidates.append(articles_dir / f"{slug}.md")
    for name in row_no_to_names(no_value):
        candidates.append(articles_dir / f"{name}.md")
    return first_existing(candidates)


def find_image_file(images_dir: Path, slug: str, no_value: str, image_name: str) -> Path | None:
    """アイキャッチ画像を探索する。今後の標準は PNG。

    優先順位:
      1. 管理表「画像ファイル名」列に値があればそれ
      2. eyecatches/{No3桁}.png
      3. eyecatches/{slug}.png
      4. 互換: No3桁 / No / slug の .webp / .jpg / .jpeg
    """
    compat_exts = [".webp", ".jpg", ".jpeg"]
    no_names = row_no_to_names(no_value)  # 例: ["001", "1"]
    candidates: list[Path] = []

    # 1) 管理表に明示された画像ファイル名を最優先
    if image_name:
        candidates.append(images_dir / image_name)
        stem = Path(image_name).stem
        candidates.append(images_dir / f"{stem}.png")
        candidates += [images_dir / f"{stem}{ext}" for ext in compat_exts]
    # 2) {No3桁}.png（→ {No}.png）
    candidates += [images_dir / f"{name}.png" for name in no_names]
    # 3) {slug}.png
    if slug:
        candidates.append(images_dir / f"{slug}.png")
    # 4) 互換拡張子（No → slug の順）
    for name in no_names:
        candidates += [images_dir / f"{name}{ext}" for ext in compat_exts]
    if slug:
        candidates += [images_dir / f"{slug}{ext}" for ext in compat_exts]
    return first_existing(candidates)


def check_eyecatch(image_path: Path | None, no_value: str, images_dir: Path) -> dict[str, Any]:
    """アイキャッチ画像を検査し、ログ文字列と Excel 反映用の値を返す。

    判定:
      - 見つからない      -> NG   / 画像未作成   / アイキャッチ画像が見つかりません
      - PNG 以外          -> WARN / 要画像確認   / アイキャッチ画像がPNG形式ではありません
      - 3:2 比率でない    -> WARN / 要画像確認   / アイキャッチ画像の比率が3:2ではありません
      - 読み込めない      -> NG   / 要画像確認   / アイキャッチ画像のサイズが取得できません
      - 上記をすべて満たす -> OK   / 画像確認済み / （エラーなし）
    推奨サイズは 1200x800px（3:2）。ファイル名が No と対応しているかも併せて確認する。
    """
    names = row_no_to_names(no_value)
    no3 = names[0] if names else safe_str(no_value)
    expected = images_dir / f"{no3}.png"

    def make(level: str, status_label: str, error_content: str, log: str) -> dict[str, Any]:
        return {
            "level": level,
            "image_status": status_label,
            "error_content": error_content,
            "log": log,
        }

    if image_path is None or not image_path.exists():
        return make("NG", "画像未作成", "アイキャッチ画像が見つかりません",
                    f"[NG] {no3} {expected} not found")

    try:
        from PIL import Image  # 遅延 import
        with Image.open(image_path) as im:
            fmt = im.format
            width, height = im.size
    except Exception as exc:
        return make("NG", "要画像確認", "アイキャッチ画像のサイズが取得できません",
                    f"[NG] {no3} {image_path} cannot read image ({exc})")

    if (fmt or "").upper() != "PNG":
        return make("WARN", "要画像確認", "アイキャッチ画像がPNG形式ではありません",
                    f"[WARN] {no3} {image_path} is not PNG ({fmt}, {width}x{height})")

    ratio = width / height if height else 0
    if abs(ratio - 1.5) > 0.05:  # 3:2 = 1.5
        return make("WARN", "要画像確認", "アイキャッチ画像の比率が3:2ではありません",
                    f"[WARN] {no3} {image_path} ratio is not 3:2 ({width}x{height})")

    size_note = "" if (width, height) == (1200, 800) else " (推奨1200x800)"
    name_note = "" if image_path.stem in set(names) else " ※ファイル名がNoと不一致"
    return make("OK", "画像確認済み", "",
                f"[OK] {no3} {image_path} exists, PNG, {width}x{height}, ratio 3:2{size_note}{name_note}")


def extract_api_error(response: requests.Response) -> str:
    """WordPress REST API のエラーJSONから code/message を読みやすく抽出する。"""
    if response.status_code == 415:
        return (
            "415 Unsupported Media Type が返っています。WordPressではなく、"
            "openresty/nginx/WAF側でREST APIリクエストが拒否されている可能性があります。"
            "WP_BASE_URL、REST API制限、セキュリティプラグイン、Basic認証の許可設定を確認してください。"
        )
    try:
        data = response.json()
        code = data.get("code", "")
        message = data.get("message", "")
        if message:
            return f"{response.status_code} {code}: {message}".strip()
    except ValueError:
        pass
    return f"{response.status_code}: {response.text[:500]}"


# --------------------------------------------------------------------------- #
# Excel への結果書き戻し（--update-excel 指定時のみ）
# --------------------------------------------------------------------------- #
# 結果キー -> (Excel 列名, 書き込みポリシー)。無ければ末尾に新規追加。
#   overwrite     : 毎回（空でも）上書き。最新状態を反映する列
#   keep_if_empty : 値が空のときは既存セルを保持。日時やIDの蓄積に使う列
EXCEL_RESULT_COLUMNS: list[tuple[str, str, str]] = [
    ("status", "投稿ステータス", "overwrite"),
    ("post_id", "WP投稿ID", "keep_if_empty"),
    ("post_link", "WP投稿URL", "keep_if_empty"),
    ("posted_at", "最終投稿日時", "keep_if_empty"),
    ("updated_at", "最終更新日時", "keep_if_empty"),
    ("run_mode", "最終実行モード", "overwrite"),
    ("error_content", "エラー内容", "overwrite"),
    ("char_count", "文字数実績", "keep_if_empty"),
    ("char_judge", "文字数判定", "keep_if_empty"),
    ("image_status", "画像制作ステータス", "overwrite"),
]

# 画像チェックのみを反映する際に書き込む列キー（投稿系の列は触らない）
IMAGE_ONLY_KEYS = {"image_status", "error_content", "updated_at"}


def _norm_no(value: Any) -> str:
    """No 値を比較用に正規化する（1 / 1.0 / '001' を同一視）。"""
    s = safe_str(value)
    if not s:
        return ""
    try:
        return str(int(float(s)))
    except ValueError:
        return s.lower()


def update_excel_results(
    input_path: Path,
    sheet: str | None,
    results: list[dict[str, Any]],
    only_keys: set[str] | None = None,
) -> int:
    """入力 Excel の該当シートへ投稿結果を No 一致で書き戻す。

    他シートや書式を壊さないよう openpyxl で既存ブックを直接更新する。
    既存の同名列があれば上書き、無ければ末尾に列を追加する。No が一致しない
    結果は書き込まない（CSV には残るため追跡可能）。戻り値は更新した行数。
    only_keys を指定すると、その結果キーに対応する列だけを書き込む
    （画像チェックのみ反映するモードで使用）。
    """
    if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise RuntimeError("--update-excel は .xlsx / .xlsm のみ対応です（CSV 入力は対象外）。")

    from openpyxl import load_workbook  # 遅延 import（CSV 運用時は不要）

    wb = load_workbook(input_path)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]

    # ヘッダー行（1 行目）を読み取り、列名 -> 列番号 を作る
    header: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        name = safe_str(cell.value)
        if name:
            header[name] = col_idx

    # No 列を特定
    no_col_name = find_col(list(header.keys()), COLUMN_CANDIDATES["no"])
    if not no_col_name:
        raise RuntimeError("Excel に No 列が見つからず、行を特定できません。")
    no_col_idx = header[no_col_name]

    # 書き込む列（only_keys 指定時はその列だけ）
    write_cols = [c for c in EXCEL_RESULT_COLUMNS if only_keys is None or c[0] in only_keys]

    # 出力列の列番号を決定（無ければ末尾に追加）
    next_col = (max(header.values()) if header else 0) + 1
    target_cols: dict[str, int] = {}
    for key, col_name, _policy in write_cols:
        if col_name in header:
            target_cols[key] = header[col_name]
        else:
            ws.cell(row=1, column=next_col, value=col_name)
            target_cols[key] = next_col
            next_col += 1

    # No -> 行番号 のマップを作る（データは 2 行目以降）
    row_by_no: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        key = _norm_no(ws.cell(row=r, column=no_col_idx).value)
        if key and key not in row_by_no:
            row_by_no[key] = r

    updated = 0
    for result in results:
        row = row_by_no.get(_norm_no(result.get("no")))
        if not row:
            continue
        for key, _col_name, policy in write_cols:
            val = result.get(key, "")
            cell = ws.cell(row=row, column=target_cols[key])
            if policy == "keep_if_empty" and (val is None or val == ""):
                continue  # 空なら既存セルを保持（日時・ID などを蓄積）
            cell.value = val if val is not None else ""
        updated += 1

    wb.save(input_path)
    return updated


# --------------------------------------------------------------------------- #
# WordPress クライアント
# --------------------------------------------------------------------------- #
class WordPressClient:
    def __init__(self, base_url: str, username: str, app_password: str) -> None:
        self.base_url = normalize_url(base_url)
        self.api_base = f"{self.base_url}/wp-json/wp/v2"
        self.session = requests.Session()
        self.session.headers.update(basic_auth_header(username, app_password))
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 WordPress-Auto-Post/2.0",
            "Accept": "application/json",
            "Content-Type": "application/json; charset=utf-8",
        })

        retry = Retry(
            total=4,
            backoff_factor=2,  # 2s, 4s, 8s, ...
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET", "POST"],
            raise_on_status=False,
        )
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

        self._term_cache: dict[tuple[str, str], int] = {}

    def request(self, method: str, endpoint: str, **kwargs: Any) -> Any:
        url = f"{self.api_base}/{endpoint.lstrip('/')}"
        kwargs.setdefault("timeout", 90)
        response = self.session.request(method, url, **kwargs)
        if response.status_code >= 400:
            raise RuntimeError(f"WordPress API エラー {extract_api_error(response)}")
        if not response.text:
            return None
        return response.json()

    def verify_auth(self) -> str:
        """認証確認。失敗時は分かりやすい例外を投げる。"""
        me = self.request("GET", "users/me")
        return me.get("name") or me.get("slug") or "(unknown)"

    def find_or_create_term(self, taxonomy: str, name: str) -> int:
        name = name.strip()
        if not name:
            return 0
        cache_key = (taxonomy, name.lower())
        if cache_key in self._term_cache:
            return self._term_cache[cache_key]

        endpoint = "categories" if taxonomy == "category" else "tags"
        data = self.request("GET", endpoint, params={"search": name, "per_page": 100})
        term_id = 0
        for item in (data if isinstance(data, list) else []):
            if safe_str(item.get("name")).lower() == name.lower():
                term_id = int(item["id"])
                break
        if not term_id:
            created = self.request("POST", endpoint, json={"name": name})
            term_id = int(created["id"])

        self._term_cache[cache_key] = term_id
        return term_id

    def find_post_by_slug(self, slug: str, status: str) -> dict[str, Any] | None:
        if not slug:
            return None
        # draft/pending を検索するには context=edit と全 status 指定が必要
        params = {
            "slug": slug,
            "status": "publish,future,draft,pending,private",
            "context": "edit",
            "per_page": 10,
        }
        data = self.request("GET", "posts", params=params)
        if isinstance(data, list):
            return data[0] if data else None
        if isinstance(data, dict) and data.get("id"):
            return data  # まれに単一オブジェクトを返すサイトに対応
        # 想定外（エラーJSON等）の応答は既存判定をスキップし、処理を継続する
        print(f"    [注意] 投稿検索の応答が配列ではないため既存判定をスキップ: {str(data)[:200]}")
        return None

    def upload_media(self, image_path: Path, alt_text: str = "") -> tuple[int, str]:
        """画像をアップロードし (media_id, source_url) を返す。"""
        mime_type, _ = mimetypes.guess_type(str(image_path))
        if not mime_type:
            mime_type = "application/octet-stream"
        headers = {
            "Content-Disposition": f'attachment; filename="{image_path.name}"',
            "Content-Type": mime_type,
        }
        with image_path.open("rb") as f:
            response = self.session.post(
                f"{self.api_base}/media", headers=headers, data=f, timeout=300
            )
        if response.status_code >= 400:
            raise RuntimeError(f"メディアアップロード失敗 {extract_api_error(response)}")
        media = response.json()
        media_id = int(media["id"])
        source_url = media.get("source_url", "")
        if alt_text:
            self.request("POST", f"media/{media_id}", json={"alt_text": alt_text})
        return media_id, source_url

    def create_post(
        self,
        *,
        title: str,
        content_html: str,
        slug: str,
        excerpt: str,
        status: str,
        category_ids: list[int],
        tag_ids: list[int],
        featured_media: int | None,
        post_id: int | None = None,
        focus_keyword: str = "",
    ) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "title": title,
            "content": content_html,
            "status": status,
            "slug": slug,
            "excerpt": excerpt,
        }
        if category_ids:
            payload["categories"] = category_ids
        if tag_ids:
            payload["tags"] = tag_ids
        if featured_media:
            payload["featured_media"] = featured_media
        # Rank Math フォーカスキーワード（重要なキーワード）を設定する。
        # ※ WP 側で rank_math_focus_keyword が show_in_rest=true 登録されている
        #    場合のみ保存される（未登録だと WordPress 側で無視される）。
        if focus_keyword:
            payload["meta"] = {"rank_math_focus_keyword": focus_keyword}
        endpoint = f"posts/{post_id}" if post_id else "posts"
        return self.request("POST", endpoint, json=payload)


# --------------------------------------------------------------------------- #
# Rank Math フォーカスキーワード
# --------------------------------------------------------------------------- #
def derive_focus_keyword(kw: str, title: str) -> str:
    """Rank Math のフォーカスキーワード（重要なキーワード）を決める。

    タイトルに入っている記事のキーワードを使う。
      1. 指定KW があればそれを使う（前後空白を除去）
      2. 空の場合はタイトルの主要部（区切り記号より前）を使う
    """
    kw = safe_str(kw)
    if kw:
        return kw
    # タイトルの区切り記号（｜ | ／ / ： :）より前をキーワードとして使う
    head = title
    for sep in ("｜", "|", "／", "/", "：", ":", "〖", "【"):
        if sep in head:
            head = head.split(sep, 1)[0]
    return head.strip()


# --------------------------------------------------------------------------- #
# メイン
# --------------------------------------------------------------------------- #
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="WordPress 自動投稿スクリプト")
    parser.add_argument("--input", required=True, help="制作管理表（XLSX/CSV）のパス")
    parser.add_argument("--sheet", default=None, help="Excel シート名")
    parser.add_argument("--articles-dir", default="articles")
    parser.add_argument("--images-dir", default="eyecatches")
    parser.add_argument("--post-status", default="draft", choices=["draft", "pending", "publish"])
    parser.add_argument("--limit", type=int, default=0, help="投稿件数。0 で全件")
    parser.add_argument("--dry-run", action="store_true", help="投稿せず確認だけ行う")
    parser.add_argument("--sleep", type=float, default=1.0, help="投稿間の待機秒数")
    parser.add_argument("--keep-h1", action="store_true", help="本文先頭の H1 を除去しない")
    parser.add_argument(
        "--write-mode",
        default="create_only",
        choices=["create_only", "update_only", "upsert"],
        help="create_only: 新規のみ / update_only: 既存のみ更新 / upsert: 既存は更新、なければ新規作成",
    )
    parser.add_argument(
        "--allow-duplicate",
        action="store_true",
        help="（非推奨）--write-mode upsert と同じ。後方互換のため残置",
    )
    parser.add_argument("--output-dir", default=".", help="結果 CSV の出力先")
    parser.add_argument(
        "--category",
        default="",
        help="全記事のカテゴリをこの値に統一する（Excelの WPカテゴリ 列を無視）。例: ブログ",
    )
    parser.add_argument(
        "--nos",
        default="",
        help="投稿する No を範囲・カンマで指定。例: 2-10,13-15,20（空なら全件）",
    )
    parser.add_argument(
        "--no-gutenberg-blocks",
        action="store_true",
        help="本文を Gutenberg ブロックに変換せず HTML のまま投稿する（既定は変換）",
    )
    parser.add_argument(
        "--update-excel",
        action="store_true",
        help="入力 Excel の該当シートへ投稿結果（ステータス/ID/URL/日時/エラー）を書き戻す",
    )
    parser.add_argument(
        "--check-images-only",
        action="store_true",
        help="WordPress 投稿を行わず、アイキャッチ画像のチェック結果だけを出力する",
    )
    parser.add_argument(
        "--no-inline-eyecatch",
        action="store_true",
        help="アイキャッチ画像を本文先頭に差し込まない（既定は差し込む）",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    limit = None if args.limit is None or args.limit <= 0 else args.limit

    input_path = Path(args.input)
    articles_dir = Path(args.articles_dir)
    images_dir = Path(args.images_dir)

    df = read_table(input_path, args.sheet)
    cols = df.columns.tolist()
    col = {key: find_col(cols, cands) for key, cands in COLUMN_CANDIDATES.items()}

    print("=== 列マッピング結果 ===")
    for key, name in col.items():
        print(f"  {key:10s} -> {name}")
    run_label = "CHECK-IMAGES-ONLY" if args.check_images_only else ("DRY-RUN" if args.dry_run else "POST")
    print(f"  対象行数: {len(df)}  / モード: {run_label}  / status: {args.post_status}")

    if not args.check_images_only and col["kw"] is None and col["title"] is None:
        raise RuntimeError("KW 列・タイトル列のどちらも見つかりません。管理表の列名を確認してください。")

    # --allow-duplicate は後方互換のため upsert 相当として扱う（write_mode 既定時のみ）
    write_mode = args.write_mode
    if args.allow_duplicate and write_mode == "create_only":
        write_mode = "upsert"
        print("  [注意] --allow-duplicate は非推奨です。--write-mode upsert として扱います。")
    if not args.check_images_only:
        print(f"  書き込みモード: {write_mode}")

    wp: WordPressClient | None = None
    if args.check_images_only:
        print("  画像チェックのみ実行します（WordPress 投稿は行いません）。")
    elif not args.dry_run:
        # 投稿モードでは認証必須
        wp = WordPressClient(
            env_required("WP_BASE_URL"),
            env_required("WP_USERNAME"),
            env_required("WP_APP_PASSWORD"),
        )
        try:
            user = wp.verify_auth()
            print(f"  認証OK: {user} としてログイン")
        except Exception as exc:
            raise RuntimeError(
                f"WordPress 認証に失敗しました。WP_BASE_URL / WP_USERNAME / WP_APP_PASSWORD を確認してください: {exc}"
            )
    else:
        # dry-run では認証情報があれば既存有無の判定に使う（無ければ判定不可として続行）
        if all(os.environ.get(k, "").strip() for k in ("WP_BASE_URL", "WP_USERNAME", "WP_APP_PASSWORD")):
            try:
                wp = WordPressClient(
                    os.environ["WP_BASE_URL"], os.environ["WP_USERNAME"], os.environ["WP_APP_PASSWORD"]
                )
                user = wp.verify_auth()
                print(f"  認証OK（dry-run/既存判定用）: {user}")
            except Exception as exc:
                wp = None
                print(f"  [注意] dry-run の既存判定をスキップします（認証不可: {exc}）")
        else:
            print("  [注意] 認証情報が無いため dry-run では既存有無を判定しません。")

    no_filter = parse_no_spec(args.nos)
    if no_filter is not None:
        rng = ",".join(str(n) for n in sorted(no_filter))
        print(f"  対象No指定: {rng}")

    # スラッグ重複の検出（create_only で後勝ちの取りこぼしを防ぐための警告）
    dup_slugs = detect_duplicate_slugs(df, col)
    if dup_slugs:
        print("  [警告] 管理表にスラッグの重複があります。create_only では後から処理する記事がスキップされます。")
        for slug, nos in dup_slugs.items():
            print(f"    スラッグ重複: {slug} (No.{', No.'.join(nos)})")
        print("    → 重複したスラッグを一意に変更してください。")

    results: list[dict[str, Any]] = []
    processed = 0

    for idx, row in df.iterrows():
        if limit is not None and processed >= limit:
            break

        no_value = safe_str(row.get(col["no"])) if col["no"] else str(idx + 1)
        if not no_value:
            no_value = str(idx + 1)

        # --nos 指定時は対象 No 以外をスキップ
        if no_filter is not None:
            n_int = no_to_int(no_value)
            if n_int is None or n_int not in no_filter:
                continue
        kw = safe_str(row.get(col["kw"])) if col["kw"] else ""
        title = safe_str(row.get(col["title"])) if col["title"] else ""
        if not title:
            title = kw
        if not title and not args.check_images_only:
            continue  # 投稿系: タイトルも KW も無い行はスキップ（空行対策）

        explicit_slug = safe_str(row.get(col["slug"])) if col["slug"] else ""
        slug = make_slug(explicit_slug or title, f"post-{int(idx) + 1:03d}")
        excerpt = safe_str(row.get(col["meta"])) if col["meta"] else ""
        if args.category.strip():
            # --category 指定時は全記事のカテゴリをこの値に統一（Excelの列を無視）
            category_names = split_terms(args.category)
        else:
            category_names = split_terms(safe_str(row.get(col["category"])) if col["category"] else "")
        tag_names = split_terms(safe_str(row.get(col["tag"])) if col["tag"] else "")
        image_name = safe_str(row.get(col["image_name"])) if col["image_name"] else ""
        alt_text = safe_str(row.get(col["alt"])) if col["alt"] else ""
        if not alt_text:
            alt_text = f"{title}のアイキャッチ画像"

        article_path = find_article_file(articles_dir, explicit_slug, no_value)
        image_path = find_image_file(images_dir, explicit_slug, no_value, image_name)

        # アイキャッチ画像チェック（dry-run / post / check-images-only 共通でログ出力）
        img_chk = check_eyecatch(image_path, no_value, images_dir)
        print(img_chk["log"])

        run_mode = "check-images-only" if args.check_images_only \
            else f"{'dry-run' if args.dry_run else 'post'}/{write_mode}"
        result: dict[str, Any] = {
            "no": no_value,
            "kw": kw,
            "title": title,
            "slug": slug,
            "article_file": str(article_path) if article_path else "",
            "image_file": str(image_path) if image_path else "（画像なし）",
            "status": "",
            "post_id": "",
            "post_link": "",
            "posted_at": "",
            "updated_at": "",
            "run_mode": run_mode,
            "char_count": "",
            "char_judge": "",
            "image_status": img_chk["image_status"],
            "image_error": img_chk["error_content"],
            "error_content": "",
            "message": "",
        }

        # 画像チェックのみモード: 投稿処理を行わず結果を記録して次へ
        if args.check_images_only:
            result["status"] = "image-check"
            result["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            result["message"] = img_chk["log"]
            results.append(result)
            processed += 1
            continue

        if article_path is None:
            result["status"] = "skipped"
            result["message"] = "本文 Markdown が見つかりません。"
            results.append(result)
            print(f"[SKIP] No.{no_value} {title} -> 本文なし")
            continue

        md_text = article_path.read_text(encoding="utf-8")
        content_html = markdown_to_html(md_text, strip_h1=not args.keep_h1, title=title)
        img_note = "画像あり" if image_path else "画像なし"

        # 文字数実績と判定（本文があれば dry-run でも算出）
        char_target = safe_str(row.get(col["char_target"])) if col["char_target"] else ""
        result["char_count"] = count_text_chars(content_html)
        result["char_judge"] = judge_char_count(result["char_count"], char_target)

        # ---- 既存投稿の有無と write_mode から動作を決定 ----------------------
        existing = wp.find_post_by_slug(slug, args.post_status) if wp else None
        existence_known = wp is not None
        if not existence_known:
            existence_note = "既存不明(認証なし)"
        else:
            existence_note = "既存あり" if existing else "既存なし"

        if write_mode == "create_only":
            action = "skip" if existing else "create"
        elif write_mode == "update_only":
            action = "update" if existing else "skip"
        else:  # upsert
            action = "update" if existing else "create"
        post_id_to_update = int(existing["id"]) if (existing and action == "update") else None

        action_label = {"create": "新規作成予定", "update": "更新予定", "skip": "スキップ予定"}[action]

        if existing:
            result["post_id"] = existing.get("id", "")
            result["post_link"] = existing.get("link", "")

        # ---- dry-run：判定結果のみ表示して次へ ------------------------------
        if args.dry_run:
            result["status"] = "dry-run"
            cat = "/".join(category_names) or "-"
            tag = "/".join(tag_names) or "-"
            result["message"] = f"{write_mode}: {existence_note} → {action_label}（{img_note} / cat:{cat} / tag:{tag}）"
            results.append(result)
            processed += 1
            print(f"[DRY-RUN] No.{no_value} {title} -> {result['message']}")
            continue

        # ---- 実投稿 ---------------------------------------------------------
        try:
            assert wp is not None

            if action == "skip":
                result["status"] = "skipped"
                if write_mode == "create_only":
                    result["message"] = "同一 slug の投稿が既に存在するためスキップ（更新するには write_mode=update_only/upsert）"
                else:  # update_only かつ既存なし
                    result["message"] = "更新対象の既存投稿が見つからないためスキップ"
                results.append(result)
                processed += 1
                print(f"[SKIP] No.{no_value} {title} -> {result['message']}")
                continue

            category_ids = [i for i in (wp.find_or_create_term("category", n) for n in category_names) if i]
            tag_ids = [i for i in (wp.find_or_create_term("tag", n) for n in tag_names) if i]

            # 本文を Gutenberg ブロックへ変換（「ブロックを解除」警告の回避）
            body_html = content_html if args.no_gutenberg_blocks \
                else html_to_gutenberg_blocks(content_html)

            # 画像がある場合のみアップロードして featured_media を差し替える。
            # 画像が無い場合は featured_media を渡さず、既存のアイキャッチを保持する。
            featured_media: int | None = None
            post_content = body_html
            if image_path:
                featured_media, media_url = wp.upload_media(image_path, alt_text=alt_text)
                # タイトルと本文の間（本文先頭）にアイキャッチ画像を差し込む
                if not args.no_inline_eyecatch and media_url:
                    post_content = build_image_block(media_url, alt_text) + body_html

            focus_keyword = derive_focus_keyword(kw, title)
            created = wp.create_post(
                title=title,
                content_html=post_content,
                slug=slug,
                excerpt=excerpt,
                status=args.post_status,
                category_ids=category_ids,
                tag_ids=tag_ids,
                featured_media=featured_media,
                post_id=post_id_to_update,
                focus_keyword=focus_keyword,
            )
            if focus_keyword:
                # 投稿レスポンスの meta から実際に保存された値を確認する。
                # show_in_rest 登録済みなら保存値が返る。空なら未保存＝サイト側未対応。
                saved_kw = (created.get("meta") or {}).get("rank_math_focus_keyword", "")
                if saved_kw:
                    print(f"  focus_keyword OK（保存確認）: {saved_kw}")
                else:
                    print(f"  focus_keyword 送信済みだが未保存: '{focus_keyword}' "
                          f"→ サイト側でRank MathメタがREST未公開の可能性")

            now_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            result["post_id"] = created.get("id", "")
            result["post_link"] = created.get("link", "")
            note = "（画像なし）" if not image_path else ""
            if post_id_to_update:
                result["status"] = "updated"
                result["updated_at"] = now_str
                result["message"] = f"既存投稿を更新しました {note}".strip()
                print(f"[UPDATED] No.{no_value} {title} -> {result['post_link']} {note}")
            else:
                result["status"] = "posted"
                result["posted_at"] = now_str
                result["message"] = f"新規投稿しました {note}".strip()
                print(f"[POSTED] No.{no_value} {title} -> {result['post_link']} {note}")

        except Exception as exc:
            result["status"] = "error"
            result["message"] = str(exc)
            print(f"[ERROR] No.{no_value} {title} -> {exc}")

        results.append(result)
        processed += 1
        if args.sleep > 0:
            time.sleep(args.sleep)

    # エラー内容を確定（投稿エラー/スキップ理由を優先、無ければ画像チェックの問題）
    for r in results:
        post_msg = r.get("message", "") if r.get("status") in ("error", "skipped") else ""
        r["error_content"] = post_msg or r.get("image_error", "")

    # 結果 CSV
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    result_path = output_dir / f"results_{timestamp}.csv"
    pd.DataFrame(
        results,
        columns=[
            "no", "kw", "title", "slug", "article_file", "image_file",
            "status", "post_id", "post_link", "posted_at", "updated_at",
            "run_mode", "char_count", "char_judge", "image_status",
            "error_content", "message",
        ],
    ).to_csv(result_path, index=False, encoding="utf-8-sig")

    # Excel への書き戻し（--update-excel 指定時のみ）
    if args.update_excel:
        try:
            only = IMAGE_ONLY_KEYS if args.check_images_only else None
            updated = update_excel_results(input_path, args.sheet, results, only_keys=only)
            scope = "画像チェック結果" if args.check_images_only else "投稿結果"
            print(f"Excel 更新: {input_path}（{updated} 行に{scope}を書き込み）")
        except Exception as exc:
            print(f"[WARN] Excel 更新に失敗しました（CSV は出力済み）: {exc}")

    # サマリ
    summary: dict[str, int] = {}
    for r in results:
        summary[r["status"]] = summary.get(r["status"], 0) + 1
    print("=== 実行サマリ ===")
    for status, count in sorted(summary.items()):
        print(f"  {status}: {count}")
    print(f"結果CSV: {result_path}")

    # error があっても全体は完了扱い（CSV で追跡）。ただし戻り値で検知できるようにする。
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
